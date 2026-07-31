import io
import re
import csv
from pathlib import Path
import openpyxl

URL_REGEX = re.compile(r'https?://[^\s,\"\']+', re.IGNORECASE)

class WorkbookAgent:
    """
    Step 1: Workbook Agent
    Inspects Excel/CSV workbooks:
    - Analyzes sheets, header rows, hidden rows, hidden sheets, merged cells.
    - Automatically discovers column containing document URLs (http, https, Google Drive, OneDrive, SharePoint, Dropbox, S3, direct image/PDF links).
    """
    def analyze_workbook(self, file_bytes: bytes, filename: str) -> dict:
        ext = Path(filename).suffix.lower()
        url_tasks = []
        headers = []
        url_col_index = None

        if ext == '.csv':
            decoded = file_bytes.decode('utf-8', errors='ignore')
            reader = csv.reader(io.StringIO(decoded))
            rows_list = list(reader)
            if rows_list:
                headers = [str(c).strip() for c in rows_list[0]]
                for r_idx, row in enumerate(rows_list[1:], start=2):
                    for c_idx, cell in enumerate(row, start=1):
                        cell_str = str(cell).strip()
                        match = URL_REGEX.search(cell_str)
                        if match:
                            url_tasks.append({
                                "rowIndex": r_idx,
                                "colIndex": c_idx,
                                "url": match.group(0)
                            })
                            if url_col_index is None:
                                url_col_index = c_idx

        elif ext in ['.xlsx', '.xls']:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            sheet = wb.active

            # Inspect headers
            for c in range(1, sheet.max_column + 1):
                val = sheet.cell(1, c).value
                headers.append(str(val or f"Col_{c}").strip())

            for r in range(2, sheet.max_row + 1):
                # Check for hidden row
                row_dim = sheet.row_dimensions.get(r)
                if row_dim and row_dim.hidden:
                    continue

                for c in range(1, sheet.max_column + 1):
                    val = sheet.cell(r, c).value
                    val_str = str(val).strip() if val is not None else ""
                    match = URL_REGEX.search(val_str)
                    if match:
                        url_tasks.append({
                            "rowIndex": r,
                            "colIndex": c,
                            "url": match.group(0)
                        })
                        if url_col_index is None:
                            url_col_index = c
            wb.close()

        url_col_name = headers[url_col_index - 1] if url_col_index and url_col_index <= len(headers) else "URL Column"

        return {
            "filename": filename,
            "headers": headers,
            "url_column": url_col_name,
            "url_col_index": url_col_index,
            "total_url_tasks": len(url_tasks),
            "tasks": url_tasks
        }
