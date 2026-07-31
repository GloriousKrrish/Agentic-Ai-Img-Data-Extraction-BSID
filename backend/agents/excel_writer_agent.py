import os
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from backend.services.data_sanitizer import normalize_field

# Enterprise Priority 20 Columns in Exact Specified Order
PRIORITY_COLUMNS = [
    {"key": "invoiceImageLink", "label": "Invoice Image Link"},
    {"key": "customerName", "label": "Customer Name"},
    {"key": "customerMobile", "label": "Customer Mobile Number"},
    {"key": "vehicleNumber", "label": "Vehicle Registration Number"},
    {"key": "invoiceNumber", "label": "Invoice Number"},
    {"key": "invoiceDate", "label": "Invoice Date"},
    {"key": "dealerName", "label": "Dealer Name"},
    {"key": "dealerGst", "label": "Dealer GST"},
    {"key": "dealerAddress", "label": "Dealer Address"},
    {"key": "tyreSize", "label": "Tyre Size"},
    {"key": "pattern", "label": "Tyre Pattern"},
    {"key": "dotCode", "label": "DOT Code"},
    {"key": "serialNumber", "label": "Serial Number"},
    {"key": "quantity", "label": "Quantity"},
    {"key": "unitCost", "label": "Unit Cost"},
    {"key": "discount", "label": "Discount"},
    {"key": "tax", "label": "Tax"},
    {"key": "grandTotal", "label": "Grand Total"},
    {"key": "confidenceScore", "label": "Confidence Score"},
    {"key": "processingStatus", "label": "Processing Status"}
]

class ExcelWriterAgent:
    """
    Step 10: Senior Analyst Resilient Excel Writer Agent (Pass 6)
    
    Generates enterprise-grade workbook:
    - 20 Priority Columns in exact order
    - Frozen header (Row 1 locked)
    - Auto-filters enabled
    - Auto column widths
    - Alternating row colors (#F8FAFC / #FFFFFF)
    - Right aligned numbers, Centered IDs/Dates/Phone/Vehicle Reg, Left aligned text
    - Clickable HYPERLINK for Source Image URLs
    - Color-coded Confidence Score cells (Green >= 85%, Yellow 70-84%, Red < 70%)
    - Resilient file lock recovery (handles Excel file lock PermissionError gracefully)
    """
    def __init__(self, output_path: str, failed_output_path: str = "failed_rows.xlsx"):
        self.output_path = output_path
        self.failed_output_path = failed_output_path
        self.failed_items = []

    def _save_workbook(self, wb, target_path: str = None) -> str:
        """Saves openpyxl workbook with automatic fallback if file is locked in Excel."""
        save_path = target_path or self.output_path
        try:
            wb.save(save_path)
            return save_path
        except PermissionError:
            base, ext = os.path.splitext(save_path)
            fallback_path = f"{base}_latest{ext}"
            print(f"\n[!] WARNING: Target file '{save_path}' is currently open in Excel or another program.")
            print(f"    Auto-redirecting output save to writeable file -> '{fallback_path}'\n")
            self.output_path = fallback_path
            wb.save(fallback_path)
            return fallback_path

    def get_effective_schema(self, schema_fields: list[dict] = None) -> list[dict]:
        """Returns Priority 20 columns schema combined with any extra discovered columns."""
        if not schema_fields:
            return PRIORITY_COLUMNS
            
        priority_keys = {c["key"] for c in PRIORITY_COLUMNS}
        extra_columns = []
        for sf in schema_fields:
            k = sf.get("key")
            if k and k not in priority_keys and k not in ["sourceUrl"]:
                extra_columns.append({"key": k, "label": sf.get("label", k.replace('_', ' ').title())})
                
        return PRIORITY_COLUMNS + extra_columns

    def init_workbook(self, schema_fields: list[dict] = None):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Enterprise Extracted Invoices"

        cols = self.get_effective_schema(schema_fields)
        headers = [c["label"] for c in cols]

        ws.append(headers)
        
        # Frozen Header A2
        ws.freeze_panes = 'A2'

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        self._save_workbook(wb)
        wb.close()

    def write_row_incremental(self, schema_fields: list[dict], row_idx: int, fields_dict: dict, confidence: float = 95.0, status: str = "COMPLETED"):
        if not os.path.exists(self.output_path):
            self.init_workbook(schema_fields)

        try:
            wb = openpyxl.load_workbook(self.output_path)
        except PermissionError:
            self.init_workbook(schema_fields)
            wb = openpyxl.load_workbook(self.output_path)

        ws = wb.active

        cols = self.get_effective_schema(schema_fields)
        field_keys = [c["key"] for c in cols]

        target_row = ws.max_row + 1
        row_fill = PatternFill(start_color="F8FAFC" if target_row % 2 == 0 else "FFFFFF", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        # Build full field map including URL, confidence, status
        full_fields = dict(fields_dict)
        if not full_fields.get("invoiceImageLink"):
            full_fields["invoiceImageLink"] = full_fields.get("sourceUrl", "")
        full_fields["confidenceScore"] = f"{confidence:.1f}%"
        full_fields["processingStatus"] = status

        for col_idx, fk in enumerate(field_keys, 1):
            raw_v = full_fields.get(fk)
            cell = ws.cell(row=target_row, column=col_idx)
            cell.fill = row_fill
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)

            fk_lower = fk.lower()

            # 1. Image Link Hyperlink
            if fk == "invoiceImageLink" or fk_lower == "sourceurl":
                url_str = str(raw_v or "").strip()
                if url_str.startswith("http://") or url_str.startswith("https://"):
                    cell.value = f'=HYPERLINK("{url_str}", "View Image")'
                    cell.font = Font(name="Calibri", size=10, color="2563EB", underline="single")
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.value = url_str
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            # 2. Confidence Score Color Coding
            elif fk == "confidenceScore":
                conf_val = float(str(raw_v).replace('%', '')) if raw_v else confidence
                cell.value = f"{conf_val:.1f}%"
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                if conf_val >= 85.0:
                    cell.fill = PatternFill(start_color="D1FAE5", fill_type="solid")
                    cell.font = Font(name="Calibri", size=10, bold=True, color="065F46") # Green
                elif conf_val >= 70.0:
                    cell.fill = PatternFill(start_color="FEF3C7", fill_type="solid")
                    cell.font = Font(name="Calibri", size=10, bold=True, color="92400E") # Yellow
                else:
                    cell.fill = PatternFill(start_color="FEE2E2", fill_type="solid")
                    cell.font = Font(name="Calibri", size=10, bold=True, color="991B1B") # Red

            # 3. Numeric Amounts Alignment & Formatting
            elif any(num_k in fk_lower for num_k in ["unitcost", "discount", "tax", "grandtotal", "cost", "price", "total", "amount"]):
                norm_v = normalize_field(fk, raw_v)
                if not norm_v:
                    cell.value = None
                else:
                    try:
                        num_val = float(norm_v)
                        cell.value = num_val
                        cell.number_format = '#,##0.00' if '.' in norm_v else '#,##0'
                    except ValueError:
                        cell.value = norm_v
                cell.alignment = Alignment(horizontal="right", vertical="center")

            # 4. Quantity Alignment
            elif "quantity" in fk_lower or fk_lower == "qty":
                norm_v = normalize_field(fk, raw_v)
                if norm_v:
                    try:
                        cell.value = int(norm_v)
                        cell.number_format = '#,##0'
                    except ValueError:
                        cell.value = norm_v
                else:
                    cell.value = None
                cell.alignment = Alignment(horizontal="right", vertical="center")

            # 5. Centered Identifiers (Date, Mobile, Reg, DOT, Serial, Status)
            elif any(id_k in fk_lower for id_k in ["number", "date", "mobile", "phone", "vehicle", "dot", "serial", "status"]):
                norm_v = normalize_field(fk, raw_v)
                cell.value = norm_v if norm_v else None
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 6. Default Text Left Alignment
            else:
                norm_v = normalize_field(fk, raw_v)
                cell.value = norm_v if norm_v else None
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Enable Auto-Filter across all columns
        ws.auto_filter.ref = ws.dimensions

        # Auto Column Width Adjustment
        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 16)

        self._save_workbook(wb)
        wb.close()

    def record_failed_row(self, row_index: int, url: str, error_msg: str):
        self.failed_items.append({
            "RowIndex": row_index,
            "URL": url,
            "Error": error_msg
        })
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Failed Extractions"
        ws.append(["Row Index", "Invoice Image Link", "Error Details"])
        
        for item in self.failed_items:
            ws.append([item["RowIndex"], item["URL"], item["Error"]])
            
        self._save_workbook(wb, self.failed_output_path)
        wb.close()
