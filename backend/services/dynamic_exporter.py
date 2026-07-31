import io
import json
import csv
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from backend.services.data_sanitizer import clean_field_value, normalize_field

def generate_dynamic_excel(extracted_items: list[dict]) -> bytes:
    """
    Creates a professionally styled Excel workbook (.xlsx) dynamically from ANY list of extracted documents.
    Formatted like a trained business analyst report:
    - Pure blank cells for missing/invalid data (zero 'null'/'None'/'N/A' strings)
    - Type-based column alignments (Numbers right, IDs center, Text left)
    - Numeric cell types for sums and formulas
    - Professional slate headers, subtle zebra rows, and auto column widths
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Intelligence Data"
    
    if not extracted_items:
        ws.append(["No Data Extracted"])
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
        
    field_keys = []
    for item in extracted_items:
        fields = item.get("fields") or item.get("extractedFields") or {}
        for k in fields.keys():
            if k not in field_keys:
                field_keys.append(k)
                
    if field_keys:
        headers = [k.replace('_', ' ').title() for k in field_keys]
    else:
        headers = ["File Name", "Document Category", "Confidence Score"]
    
    # 1. Header Styling
    ws.append(headers)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    # 2. Append Data Rows with Professional Alignment & Numeric Types
    for row_idx, item in enumerate(extracted_items, 2):
        fields = item.get("fields") or item.get("extractedFields") or {}
        row_fill = PatternFill(start_color="F8FAFC" if row_idx % 2 == 0 else "FFFFFF", fill_type="solid")
        
        if field_keys:
            for col_idx, fk in enumerate(field_keys, 1):
                raw_v = fields.get(fk)
                norm_v = normalize_field(fk, raw_v)
                cell = ws.cell(row=row_idx, column=col_idx)
                
                if not norm_v:
                    cell.value = None
                else:
                    # Attempt numeric conversion for sums/financial metrics
                    fk_lower = fk.lower()
                    if any(num_k in fk_lower for num_k in ["amount", "cost", "price", "total", "salary", "revenue", "quantity", "units"]):
                        try:
                            if '.' in norm_v:
                                cell.value = float(norm_v)
                                cell.number_format = '#,##0.00'
                            else:
                                cell.value = int(norm_v)
                                cell.number_format = '#,##0'
                        except ValueError:
                            cell.value = norm_v
                    else:
                        cell.value = norm_v

                cell.fill = row_fill
                cell.border = thin_border
                cell.font = Font(name="Calibri", size=10)
                
                # Determine Alignment by Data Type & Field Key
                fk_lower = fk.lower()
                if any(num_k in fk_lower for num_k in ["amount", "cost", "price", "total", "salary", "revenue", "quantity", "units"]):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif any(id_k in fk_lower for id_k in ["number", "id", "code", "mobile", "phone", "date", "vehicle"]):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            c1 = ws.cell(row=row_idx, column=1, value=clean_field_value(item.get("fileName", f"Document_{row_idx-1}")))
            c2 = ws.cell(row=row_idx, column=2, value=clean_field_value(item.get("category", "General Document")))
            c3 = ws.cell(row=row_idx, column=3, value=f"{item.get('confidence', 95.0)}%")
            for c in (c1, c2, c3):
                c.fill = row_fill
                c.border = thin_border
                c.font = Font(name="Calibri", size=10)
            c1.alignment = Alignment(horizontal="left", vertical="center")
            c2.alignment = Alignment(horizontal="left", vertical="center")
            c3.alignment = Alignment(horizontal="center", vertical="center")
            
    # Auto-adjust column widths cleanly
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 14)
        
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

def generate_dynamic_csv(extracted_items: list[dict]) -> str:
    """Generates clean CSV text content dynamically."""
    if not extracted_items:
        return "File Name,Category,Status\n"
        
    field_keys = []
    for item in extracted_items:
        fields = item.get("fields") or item.get("extractedFields") or {}
        for k in fields.keys():
            if k not in field_keys:
                field_keys.append(k)
                
    output = io.StringIO()
    writer = csv.writer(output)
    
    if field_keys:
        headers = [k.replace('_', ' ').title() for k in field_keys]
        writer.writerow(headers)
        for item in extracted_items:
            fields = item.get("fields") or item.get("extractedFields") or {}
            row = [normalize_field(fk, fields.get(fk)) for fk in field_keys]
            writer.writerow(row)
    else:
        headers = ["File Name", "Category", "Confidence"]
        writer.writerow(headers)
        for item in extracted_items:
            writer.writerow([
                clean_field_value(item.get("fileName", "")),
                clean_field_value(item.get("category", "")),
                item.get("confidence", 95.0)
            ])
            
    return output.getvalue()
