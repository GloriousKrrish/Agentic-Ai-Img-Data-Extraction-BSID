import io
import os
import re
import csv
import json
import time
import requests
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PIL import Image, ImageEnhance, ImageOps

from backend.services.schema_generator import generate_dynamic_schema
from backend.services.universal_extractor import extract_universal_document

# =====================================================================
# AGENT 1: WORKBOOK ANALYZER
# =====================================================================
class WorkbookAnalyzerAgent:
    def analyze(self, file_bytes: bytes, filename: str) -> dict:
        ext = Path(filename).suffix.lower()
        url_pattern = re.compile(r'https?://[^\s,\"\']+')
        url_cells = []
        headers = []

        if ext == '.csv':
            decoded = file_bytes.decode('utf-8', errors='ignore')
            reader = csv.reader(io.StringIO(decoded))
            for row_idx, row in enumerate(reader, start=1):
                if row_idx == 1:
                    headers = [str(c).strip() for c in row]
                for col_idx, cell in enumerate(row, start=1):
                    val_str = str(cell).strip()
                    if url_pattern.search(val_str):
                        url_cells.append({"row": row_idx, "col": col_idx, "url": url_pattern.search(val_str).group(0)})
        elif ext in ['.xlsx', '.xls']:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            sheet = wb.active
            for c in range(1, sheet.max_column + 1):
                h_val = sheet.cell(1, c).value
                headers.append(str(h_val or f"Col_{c}").strip())

            for r in range(1, sheet.max_row + 1):
                for c in range(1, sheet.max_column + 1):
                    val = sheet.cell(r, c).value
                    val_str = str(val).strip() if val is not None else ""
                    match = url_pattern.search(val_str)
                    if match:
                        url_cells.append({"row": r, "col": c, "url": match.group(0)})
            wb.close()

        return {
            "total_rows": len(url_cells),
            "headers": headers,
            "url_tasks": url_cells
        }

# =====================================================================
# AGENT 2: URL VALIDATOR
# =====================================================================
class URLValidatorAgent:
    def validate(self, url: str) -> dict:
        if not url.startswith("http://") and not url.startswith("https://"):
            return {"valid": False, "reason": "Invalid URL protocol"}
        return {"valid": True, "url": url}

# =====================================================================
# AGENT 3: DOCUMENT FETCHER
# =====================================================================
class DocumentFetcherAgent:
    def fetch(self, url: str, max_retries: int = 3) -> dict:
        for attempt in range(max_retries):
            try:
                res = requests.get(url, timeout=15)
                if res.status_code == 200 and len(res.content) > 100:
                    content_type = res.headers.get("Content-Type", "image/jpeg").split(";")[0].strip().lower()
                    mime_type = "application/pdf" if ("pdf" in url.lower() or "pdf" in content_type) else "image/jpeg"
                    return {"success": True, "bytes": res.content, "mime_type": mime_type}
            except Exception as e:
                if attempt == max_retries - 1:
                    return {"success": False, "error": str(e)}
                time.sleep(1.0)
        return {"success": False, "error": "HTTP Download Failed"}

# =====================================================================
# AGENT 4: IMAGE PREPROCESSOR
# =====================================================================
class ImagePreprocessorAgent:
    def preprocess(self, doc_bytes: bytes, mime_type: str) -> bytes:
        if mime_type == "application/pdf":
            return doc_bytes
        try:
            img = Image.open(io.BytesIO(doc_bytes))
            img = ImageOps.exif_transpose(img)
            if img.mode != 'RGB':
                img = img.convert('RGB')
            enhancer = ImageEnhance.Contrast(img)
            img = enhancer.enhance(1.2)
            buf = io.BytesIO()
            img.save(buf, format='JPEG', quality=95)
            return buf.getvalue()
        except Exception:
            return doc_bytes

# =====================================================================
# AGENT 5: CLASSIFIER & AGENT 6/7/8: SEMANTIC EXTRACTION AGENTS
# =====================================================================
SEMANTIC_FULL_SCHEMA = {
    "documentCategory": "Invoice / Transaction Document",
    "documentTitle": "Intelligent Extracted Invoice Data",
    "summary": "Extracted semantic key-value fields from document",
    "fields": [
        {"key": "invoiceNumber", "label": "Invoice Number", "type": "string", "description": "Invoice number or bill reference"},
        {"key": "invoiceDate", "label": "Invoice Date", "type": "string", "description": "Invoice issue date"},
        {"key": "customerName", "label": "Customer Name", "type": "string", "description": "Customer full name"},
        {"key": "customerMobile", "label": "Customer Mobile", "type": "string", "description": "Customer phone number"},
        {"key": "vehicleNumber", "label": "Vehicle Number", "type": "string", "description": "Vehicle registration number (e.g. AP39NT1461)"},
        {"key": "vehicleModel", "label": "Vehicle Model", "type": "string", "description": "Vehicle model name"},
        {"key": "tyreSize", "label": "Tyre Size", "type": "string", "description": "Tyre size specification (e.g. 235/65R17)"},
        {"key": "pattern", "label": "Pattern", "type": "string", "description": "Tyre pattern"},
        {"key": "serialNumber", "label": "Serial Number", "type": "string", "description": "Product serial number"},
        {"key": "dealerName", "label": "Dealer Name", "type": "string", "description": "Dealer or shop name"},
        {"key": "price", "label": "Price / Total", "type": "number", "description": "Total price or amount"},
        {"key": "remarks", "label": "Remarks", "type": "string", "description": "Additional notes"}
    ]
}

class ExtractionAgent:
    def extract(self, doc_bytes: bytes, mime_type: str) -> dict:
        try:
            schema_info = generate_dynamic_schema(doc_bytes, mime_type)
            ext_res = extract_universal_document(doc_bytes, schema_info, mime_type)
            rows = ext_res.get("rows", [])
            if rows and len(rows) > 0:
                return {
                    "category": schema_info.get("documentCategory", "Invoice Document"),
                    "fields": rows[0].get("fields", {})
                }
        except Exception:
            pass

        # Fallback to semantic schema
        try:
            ext_res = extract_universal_document(doc_bytes, SEMANTIC_FULL_SCHEMA, mime_type)
            rows = ext_res.get("rows", [])
            if rows and len(rows) > 0:
                return {
                    "category": "Invoice Document",
                    "fields": rows[0].get("fields", {})
                }
        except Exception:
            pass

        return {"category": "General Document", "fields": {}}

# =====================================================================
# AGENT 9: SCHEMA EVOLUTION AGENT
# =====================================================================
class SchemaEvolutionAgent:
    def __init__(self):
        self.field_keys = set()

    def register(self, fields: dict) -> list[str]:
        for k in fields.keys():
            self.field_keys.add(k)
        return sorted(list(self.field_keys))

# =====================================================================
# AGENT 10: VALIDATION AGENT
# =====================================================================
class ValidationAgent:
    def validate(self, fields: dict) -> dict:
        confidence = 95.0
        warnings = []
        if not fields.get("invoiceNumber") and not fields.get("vehicleNumber"):
            warnings.append("Missing primary reference number")
            confidence -= 10.0
        return {"confidence": max(confidence, 70.0), "warnings": warnings}

# =====================================================================
# AGENT 11: RESILIENT EXCEL WRITER
# =====================================================================
from backend.services.data_sanitizer import normalize_field

class ResilientExcelWriterAgent:
    def write_workbook(self, items: list[dict], output_path: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extracted Intelligence Data"

        if not items:
            ws.append(["No Data Extracted"])
            wb.save(output_path)
            return

        all_keys = set()
        for it in items:
            for k in it.get("fields", {}).keys():
                all_keys.add(k)

        field_keys = sorted(list(all_keys))
        headers = [k.replace('_', ' ').title() for k in field_keys]
        ws.append(headers)

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, it in enumerate(items, 2):
            f_dict = it.get("fields", {})
            row_fill = PatternFill(start_color="F8FAFC" if row_idx % 2 == 0 else "FFFFFF", fill_type="solid")
            for col_idx, fk in enumerate(field_keys, 1):
                raw_v = f_dict.get(fk)
                norm_v = normalize_field(fk, raw_v)
                cell = ws.cell(row=row_idx, column=col_idx)
                
                if not norm_v:
                    cell.value = None
                else:
                    fk_lower = fk.lower()
                    if any(num_k in fk_lower for num_k in ["amount", "cost", "price", "total", "salary", "revenue", "quantity", "units"]):
                        try:
                            if '.' in norm_v:
                                cell.value = float(norm_v)
                                cell.number_format = '#,##0.00'
                            else:
                                cell.value = int(norm_v)
                                cell.number_format = '#,##0'
                        except ValueError:
                            cell.value = norm_v
                    else:
                        cell.value = norm_v

                cell.fill = row_fill
                cell.border = thin_border
                cell.font = Font(name="Calibri", size=10)
                
                fk_lower = fk.lower()
                if any(num_k in fk_lower for num_k in ["amount", "cost", "price", "total", "salary", "revenue", "quantity"]):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif any(id_k in fk_lower for id_k in ["number", "id", "code", "mobile", "phone", "date", "vehicle"]):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # Auto column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 14)

        tmp_path = f"{output_path}.tmp"
        try:
            wb.save(tmp_path)
            if os.path.exists(output_path):
                os.remove(output_path)
            os.replace(tmp_path, output_path)
        except Exception:
            try:
                wb.save(output_path)
            except Exception:
                pass
