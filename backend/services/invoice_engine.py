import io
import re
import csv
import requests
from pathlib import Path
import openpyxl
from backend.services.universal_extractor import extract_universal_document
from backend.services.data_sanitizer import sanitize_extracted_dict

# 20 Priority Columns Standardized Enterprise Schema
INVOICE_SEMANTIC_PROMPT_SCHEMA = {
    "documentCategory": "Invoice / Transaction Document",
    "documentTitle": "Enterprise Cognitive Invoice Data Extraction",
    "summary": "Extracted semantic key-value fields from invoice document with business understanding",
    "fields": [
        {"key": "customerName", "label": "Customer Name", "type": "string", "description": "Full name of buyer or customer"},
        {"key": "customerMobile", "label": "Customer Mobile Number", "type": "string", "description": "10-digit mobile number starting with 6-9"},
        {"key": "vehicleNumber", "label": "Vehicle Registration Number", "type": "string", "description": "Vehicle registration number (e.g. AP39NT1461)"},
        {"key": "invoiceNumber", "label": "Invoice Number", "type": "string", "description": "Invoice number or bill reference"},
        {"key": "invoiceDate", "label": "Invoice Date", "type": "string", "description": "Date invoice was issued (DD/MM/YYYY or YYYY-MM-DD)"},
        {"key": "dealerName", "label": "Dealer Name", "type": "string", "description": "Name of dealership, shop, or company issuing invoice"},
        {"key": "dealerGst", "label": "Dealer GST", "type": "string", "description": "15-character GSTIN number of dealer or seller"},
        {"key": "dealerAddress", "label": "Dealer Address", "type": "string", "description": "Full address of dealer"},
        {"key": "tyreSize", "label": "Tyre Size", "type": "string", "description": "Tyre specification size code (e.g. 235/65R17)"},
        {"key": "pattern", "label": "Tyre Pattern", "type": "string", "description": "Tyre pattern or tread design name"},
        {"key": "dotCode", "label": "DOT Code", "type": "string", "description": "DOT manufacturing batch code (e.g. DOT 4223)"},
        {"key": "serialNumber", "label": "Serial Number", "type": "string", "description": "Product or tyre serial number"},
        {"key": "quantity", "label": "Quantity", "type": "number", "description": "Number of units purchased"},
        {"key": "unitCost", "label": "Unit Cost", "type": "number", "description": "Price per single unit"},
        {"key": "discount", "label": "Discount", "type": "number", "description": "Discount amount applied"},
        {"key": "tax", "label": "Tax", "type": "number", "description": "Total GST or tax amount"},
        {"key": "grandTotal", "label": "Grand Total", "type": "number", "description": "Final total payable amount"}
    ]
}

def detect_url_column_in_excel(file_bytes: bytes, filename: str) -> list[dict]:
    """
    Scans any Excel or CSV workbook and automatically extracts HTTP/HTTPS URLs from any column or row.
    Returns: [{"rowIndex": 2, "url": "https://..."}]
    """
    ext = Path(filename).suffix.lower()
    urls = []
    url_pattern = re.compile(r'https?://[^\s,\"\']+')

    try:
        if ext == '.csv':
            decoded = file_bytes.decode('utf-8', errors='ignore')
            reader = csv.reader(io.StringIO(decoded))
            for row_idx, row in enumerate(reader, start=1):
                for cell in row:
                    cell_str = str(cell).strip()
                    match = url_pattern.search(cell_str)
                    if match:
                        urls.append({"rowIndex": row_idx, "url": match.group(0)})
        elif ext in ['.xlsx', '.xls']:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            sheet = wb.active
            for row_idx in range(1, sheet.max_row + 1):
                for col_idx in range(1, sheet.max_column + 1):
                    val = sheet.cell(row=row_idx, column=col_idx).value
                    val_str = str(val).strip() if val is not None else ""
                    match = url_pattern.search(val_str)
                    if match:
                        urls.append({"rowIndex": row_idx, "url": match.group(0)})
            wb.close()
    except Exception as e:
        print(f"Error scanning workbook for URLs: {e}")

    return urls

def process_invoice_document(doc_bytes: bytes, mime_type: str = "image/jpeg", text_content: str = "") -> dict:
    """
    Processes an invoice document through Vision AI and extracts priority semantic fields.
    """
    try:
        ext_res = extract_universal_document(
            doc_bytes, 
            INVOICE_SEMANTIC_PROMPT_SCHEMA, 
            mime_type,
            text_content=text_content
        )
        rows = ext_res.get("rows", [])
        if rows and len(rows) > 0:
            fields = rows[0].get("fields", {})
            return sanitize_extracted_dict(fields, min_confidence=60.0)
        return {}
    except Exception as e:
        print(f"Error processing invoice document: {e}")
        return {}
