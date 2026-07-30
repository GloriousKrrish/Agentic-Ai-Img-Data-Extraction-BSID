import json
import re
from pathlib import Path
import openpyxl
from backend.config import EXCEL_PATH, PROJECT_ENGINE_DIR

def slugify(text: str) -> str:
    """Converts column title text into camelCase or safe key string."""
    clean = re.sub(r'[^a-zA-Z0-9\s]', '', str(text or '')).strip()
    words = clean.split()
    if not words:
        return "col"
    return words[0].lower() + "".join(w.capitalize() for w in words[1:])

def read_excel_rows(file_path: Path = EXCEL_PATH) -> dict:
    """
    Dynamically reads ANY Excel spreadsheet file.
    Row 1 is automatically parsed as column headers.
    Returns:
    {
       "schema": [{"key": "colKey", "label": "Column Label"}],
       "rows": [{"rowIndex": 2, "fields": {"colKey": "val"}, "status": "COMPLETED"}]
    }
    """
    if not file_path.exists():
        return {"schema": [], "rows": []}
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        
        # 1. Dynamically read headers from Row 1
        headers = []
        schema = []
        for col_idx in range(1, sheet.max_column + 1):
            val = sheet.cell(row=1, column=col_idx).value
            if val is not None and str(val).strip():
                label = str(val).strip()
                key = slugify(label)
                # ensure uniqueness of keys
                if any(s["key"] == key for s in schema):
                    key = f"{key}_{col_idx}"
                schema.append({"key": key, "label": label, "colIndex": col_idx})
                
        if not schema:
            wb.close()
            return {"schema": [], "rows": []}
            
        # 2. Dynamically read rows starting from Row 2
        rows = []
        for row_idx in range(2, sheet.max_row + 1):
            row_fields = {}
            has_data = False
            for col in schema:
                cell_val = sheet.cell(row=row_idx, column=col["colIndex"]).value
                str_val = str(cell_val).strip() if cell_val is not None else ""
                row_fields[col["key"]] = str_val
                if str_val and not str_val.startswith("http"):
                    has_data = True
                    
            rows.append({
                "rowIndex": row_idx,
                "fields": row_fields,
                "status": "COMPLETED" if has_data else "PENDING"
            })
            
        wb.close()
        # Clean schema output (remove internal colIndex)
        clean_schema = [{"key": s["key"], "label": s["label"]} for s in schema]
        return {
            "schema": clean_schema,
            "rows": rows
        }
    except Exception as e:
        return {"schema": [], "rows": [], "error": str(e)}

def extract_urls_from_excel_bytes(file_bytes: bytes, filename: str) -> list[dict]:
    """
    Parses Excel/CSV file bytes and extracts all image HTTP/HTTPS URL links.
    Returns: [{"rowIndex": 2, "url": "https://..."}]
    """
    ext = Path(filename).suffix.lower()
    urls_list = []

    try:
        if ext == '.csv':
            decoded = file_bytes.decode('utf-8', errors='ignore')
            reader = csv.reader(io.StringIO(decoded))
            for row_idx, row in enumerate(reader, start=1):
                for cell in row:
                    cell_str = str(cell).strip()
                    if cell_str.startswith("http://") or cell_str.startswith("https://"):
                        urls_list.append({"rowIndex": row_idx, "url": cell_str})
        elif ext in ['.xlsx', '.xls']:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            sheet = wb.active
            for row_idx in range(1, sheet.max_row + 1):
                for col_idx in range(1, sheet.max_column + 1):
                    val = sheet.cell(row=row_idx, column=col_idx).value
                    val_str = str(val).strip() if val is not None else ""
                    if val_str.startswith("http://") or val_str.startswith("https://"):
                        urls_list.append({"rowIndex": row_idx, "url": val_str})
            wb.close()
    except Exception as e:
        print(f"Error extracting URLs from Excel: {e}")

    return urls_list

def get_excel_kpis(file_path: Path = EXCEL_PATH):
    result = read_excel_rows(file_path)
    rows = result.get("rows", [])
    total_rows = len(rows)
    processed = [r for r in rows if r["status"] == "COMPLETED"]
    processed_count = len(processed)
    pending_count = total_rows - processed_count
    
    return {
        "totalDocuments": total_rows,
        "processedDocuments": processed_count,
        "pendingDocuments": pending_count,
        "successRate": round((processed_count / total_rows * 100), 1) if total_rows > 0 else 0.0
    }
