import openpyxl
import requests
import io
import json
from pathlib import Path
from backend.services.schema_generator import generate_dynamic_schema
from backend.services.universal_extractor import extract_universal_document
from backend.services.dynamic_exporter import generate_dynamic_excel

def extract_all_testing_rows():
    excel_path = Path("testing.xlsx")
    if not excel_path.exists():
        print("[FAIL] testing.xlsx not found!")
        return

    print("=========================================================================")
    print("EXTRACTING ALL 10 IMAGE LINKS FROM testing.xlsx WITH AI VISION")
    print("=========================================================================\n")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb.active

    urls = []
    for r in range(2, sheet.max_row + 1):
        val = sheet.cell(r, 1).value
        if val and str(val).startswith("http"):
            urls.append((r - 1, str(val).strip()))
    wb.close()

    print(f"[OK] Found {len(urls)} image links in testing.xlsx.")

    all_rows = []
    common_schema = []
    common_fields = []

    for idx, (row_num, url) in enumerate(urls, start=1):
        print(f"[{idx}/{len(urls)}] Fetching & Extracting Image from: {url[:60]}...")
        try:
            res = requests.get(url, timeout=15)
            if res.status_code == 200 and len(res.content) > 100:
                img_bytes = res.content
                if not common_schema:
                    print("  -> Inferring dynamic schema with Gemini Vision AI...")
                    schema_info = generate_dynamic_schema(img_bytes, "image/jpeg")
                    common_fields = schema_info.get("fields", [])
                else:
                    schema_info = {"documentCategory": "Invoice Document", "fields": common_fields}

                print("  -> Extracting structured fields...")
                ext_res = extract_universal_document(img_bytes, schema_info, "image/jpeg")
                for r in ext_res.get("rows", []):
                    r["rowIndex"] = row_num
                    r["fields"]["imageUrl"] = url
                    all_rows.append(r)
                    print(f"  [SUCCESS Row #{row_num}]: {r.get('fields')}")
        except Exception as e:
            print(f"  [ERROR Row #{row_num}]: {e}")

    print("\n=========================================================================")
    print(f"SUMMARY: Successfully Extracted {len(all_rows)} Rows from testing.xlsx")
    print("=========================================================================\n")

    # Generate and save new Excel file
    excel_bytes = generate_dynamic_excel(all_rows)
    out_file = "extracted_testing_results_v2.xlsx"
    with open(out_file, "wb") as f_out:
        f_out.write(excel_bytes)

    print(f"[SUCCESS] Saved extracted data into brand new Excel file: '{out_file}' ({len(excel_bytes)} bytes)")

if __name__ == "__main__":
    extract_all_testing_rows()
